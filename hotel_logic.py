import pandas as pd
import json

from openpyxl import load_workbook

# account_sid = "YOUR_SID_HERE"
# auth_token = "YOUR_TOKEN_HERE"
twilio_number = "+19027018815"

# roomSchedule = pd.ExcelFile('pages/room_schedule_template.xlsx')
# df = roomSchedule.parse('roomNum')
# rl = roomSchedule.parse('roomList')
# wb = load_workbook('pages/room_schedule.xlsx')
# ws = wb['roomList']
# print(roomSchedule)

# open excel file and read roomNum page
def schedule_rn(file):
    roomSchedule = pd.ExcelFile(file)
    rn = roomSchedule.parse('roomNum')
    return rn
# print(schedule_rn(roomSchedule))
# open excel file and read roomList page
def schedule_rl(file):
    roomSchedule = pd.ExcelFile(file)
    rl = roomSchedule.parse('roomList')
    return rl

# open excel file and read roomList page by workbook function
def schedule_ws(file):
    wb = load_workbook(file)
    ws = wb['roomList']
    return ws

# function counting the number of input room in schedule
def roomCount(roomNum, file):
    room_count = 0
    for row in schedule_ws(file).iter_rows():
        for cell in row:
            if cell.value == roomNum:
                room_count += 1
    return room_count

# testing function
def hello():
    return "Hello, this is Alex's summer project"

def sentSuccess(text):
    return f"sentSuccess: {text}"

# adding json data
def update_json(new_data, style, filename='data.json'):
    with open(filename, 'r+') as file:
        # Load existing data into a dictionary
        file_data = json.load(file)        
        # Append new data to the 'details' list
        file_data["message"].insert(0, new_data)
        file_data["type"].insert(0, style)
        # Move the cursor to the beginning of the original file
        file.seek(0)
        # delete all content after cursor
        file.truncate()       
        # Write the updated data back to the original file
        json.dump(file_data, file, indent=4)
    





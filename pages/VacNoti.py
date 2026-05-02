import streamlit as st
import hotel_logic
import openpyxl as xl
from twilio.rest import Client

if not st.session_state.get("logged_in", False):
    st.error("Please login first!")
    st.switch_page("main.py")
    st.stop()

# st.balloons()
st.image("logo.png", width=400)
st.title(hotel_logic.hello())
st.title("🏨 Room Dispatcher")
st.write("Welcome, Supervisor. Select a room to notify an attendant.")

schedule = st.file_uploader("Please upload today schedule", type=".xlsx", accept_multiple_files=False, 
                            max_upload_size=10)

if schedule is not None:
    # hotel_logic.schedule_rn(schedule)
    with st.form("notification"):
        room_number = st.text_input("Enter Room Number:", placeholder="room number")
        send = st.form_submit_button("Send")
        if send:
            try:
                int_room_number = int(room_number)
                if int_room_number in hotel_logic.schedule_rn(schedule).values:
                    # st.success(f"Notification {room_number}!")
                    if int_room_number in hotel_logic.schedule_rl(schedule).values:
                        room_count = 0
                        for row in hotel_logic.schedule_ws(schedule).iter_rows():
                            for cell in row:
                                if cell.value == int_room_number:
                                    room_count += 1
                        if room_count == 1:
                            for row in hotel_logic.schedule_ws(schedule).iter_rows():
                                for cell in row:
                                    if cell.value == int_room_number:                                      
                                        st.success(f"{hotel_logic.currTime} Notification sent to {hotel_logic.schedule_ws(schedule).cell(row=cell.row, column=1).value} \
                                    {hotel_logic.schedule_ws(schedule).cell(row=cell.row, column=2).value} for Room {room_number}!")
                        else:
                            st.error(f"{hotel_logic.currTime} The room {room_number} is assigned to more than one room attendant.")
                    else:
                        st.error(f"{hotel_logic.currTime} The room {room_number} is not assigned yet.")
                        # st.success(f"Notification sent to for Room {room_number}!")
                        
                else:
                    # error if input invalid room number
                    st.error(f"{hotel_logic.currTime} Please input a valid room number.")

            # error if input other than integer
            except ValueError:
                st.error(f"{hotel_logic.currTime} Please input a valid room number.")






    # SMS senting part
    # account_sid = functions.account_sid
    # account_token = functions.auth_token

    # client = Client(account_sid, account_token)

    # message = client.messages.create(
    #     from_=functions.twilio_number,
    #     body= f"From supervisor. Hi, room {room_number} is vacant now, thank you",
    #     to=functions.my_phone_number
    # )
    # print(message.sid)

